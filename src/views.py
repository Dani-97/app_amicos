from kivy.app import App
from kivy.core.text import LabelBase
from kivy.core.window import Window
from kivy.lang import Builder
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.screenmanager import Screen, ScreenManager
from KivyCustom.Custom import ButtonRnd, CustomTextInput
from KivyCustom.Custom import Tile
from kivy.uix.scrollview import ScrollView
from openpyxl import load_workbook

from utils_exceptions import LogInException
from utils_exceptions import InvalidUsernameException
from utils_exceptions import InvalidEmailException
from utils_exceptions import InvalidPasswordException
from utils_exceptions import DuplicatedUserException
from utils_exceptions import UserDeletionException

class AmicosApp(App):

    def set_ui_assets_manager(self, ui_assets_manager):
        self.ui_assets_manager = ui_assets_manager

    def set_message_builder_provider(self, message_builder_provider):
        self.message_builder_provider = message_builder_provider

    def set_profiles_manager_provider(self, profiles_manager_provider):
        self.profiles_manager_provider = profiles_manager_provider

    def set_settings_provider(self, settings_provider):
        self.settings_provider = settings_provider

    """
    Clase principal de la aplicación, se encarga de la inicialización
    Es la encargada de iniciar la aplicación y construirla.    
    """
    def build(self):
        self.title = 'App Amicos'
        
        self.icon = self.ui_assets_manager.get_resource("assets/logo.png")
        
        LabelBase.register(name='Titulo', fn_regular=self.ui_assets_manager.get_resource('src/fonts/Orbitron-Regular.ttf'))
        LabelBase.register(name='Texto', fn_regular=self.ui_assets_manager.get_resource('src/fonts/FrancoisOne-Regular.ttf'))

        Builder.load_file(self.ui_assets_manager.get_resource('src/KivyCustom/custom.kv'))
        main_window = MainWindow(self.ui_assets_manager,
                                 self.message_builder_provider,
                                 self.profiles_manager_provider,
                                 self.settings_provider,
                                 name="main_window")
        main_window.build()

        Builder.load_file("src/signupapp.kv")
        signup_screen = SignupScreen(self.profiles_manager_provider,
                                     name="signup_screen")
        Builder.load_file("src/loginapp.kv")
        login_screen = LogInScreen(self.profiles_manager_provider,
                                   name="login_screen")

        self.sm = ScreenManager()
        self.sm.add_widget(signup_screen)
        self.sm.add_widget(login_screen)
        self.sm.add_widget(main_window)

        current_user = self.profiles_manager_provider.get_current_user()

        if (current_user!=None):
            self.sm.current = "main_window"
        else:
            self.sm.current = "signup_screen"

        return self.sm

class GridWindow(GridLayout):

    """
    Esta clase crea un tablero con las palabras y las imágenes correspondientes
    Patron de diseño: Composite
    """
    def __init__(self, parent_window, tab_id, **kwargs):
        super(GridWindow, self).__init__(**kwargs)
        self.tab_id = tab_id
        self.parent_window = parent_window

    def set_ui_assets_manager(self, ui_assets_manager):
        self.ui_assets_manager = ui_assets_manager

    def set_content(self, words_with_images, tile_mode):
        self.words_with_images = words_with_images
        self.tile_mode = tile_mode

    def set_message_builder_provider(self, message_builder_provider):
        self.message_builder_provider = message_builder_provider

    def build(self):
        self.rows = len(self.words_with_images)
        self.cols = len(self.words_with_images[0]) if self.words_with_images else 0
        self.spacing = [10, 10]
        self.tiles = []
        for row in self.words_with_images:
            for picture, word in row:
                tile = Tile(str(word),
                            self.ui_assets_manager.get_resource(f'src/assets/{picture}'),
                            self.tile_mode,
                            on_press=self.on_tile_clicked)
                self.tiles.append(tile)
                self.add_widget(tile)

    # Función que se ejecuta al pulsar una casilla
    def on_tile_clicked(self, instance):
        tile_text = instance.tile_associated_text
        
        try:        
            self.parent_window.swap_grid_window(tile_text)
        except KeyError:
            self.message_builder_provider.add_word_to_message(tile_text)
            self.parent_window.swap_grid_window(self.tab_id)

class LogInScreen(Screen):

    def __init__(self, profiles_manager_provider, **kwargs):
        super(LogInScreen, self).__init__(**kwargs)
        self.profiles_manager_provider = profiles_manager_provider

    def go_to_signup_page(self):
        self.manager.current = 'signup_screen'  # Switch to the login screen

    def go_to_main_window(self):
        self.manager.current = 'main_window'

    def build(self):
        pass

    def validate_login(self):
        username = self.ids.username.text
        password = self.ids.password.text

        try:
            _ = self.profiles_manager_provider.login_user(username, password)
            self.ids.message.text = f"[color=00ff00]Inicio de sesión correcto[/color]"
            self.go_to_main_window()
        except LogInException:
            self.ids.message.text = "[color=ff0000]Nombre de usuario o contraseña incorrectos[/color]"

class SignupScreen(Screen):

    def __init__(self, profiles_manager_provider, **kwargs):
        super(SignupScreen, self).__init__(**kwargs)
        Window.size = (2048, 1080)
        self.profiles_manager_provider = profiles_manager_provider

    def go_to_login_page(self):
        self.manager.current = 'login_screen'  # Switch to the login screen

    def go_to_main_window(self):
        self.manager.current = 'main_window'  # Switch to the login screen

    def execute_signup(self):        
        username = self.ids.username.text
        email = self.ids.email.text
        role = self.ids.role.text
        password = self.ids.password.text

        try: 
            self.profiles_manager_provider.create_user(username, email, role, password)
            self.ids.message.text = f"[color=00ff00]Cuenta creada correctamente[/color]"
            _ = self.profiles_manager_provider.login_user(username, password)
            self.go_to_main_window()
        except InvalidUsernameException as except_obj:
            self.ids.message.text = f"[color=ff0000]{except_obj}[/color]"
        except InvalidEmailException as except_obj:
            self.ids.message.text = f"[color=ff0000]{except_obj}[/color]"
        except InvalidPasswordException as except_obj:
            self.ids.message.text = f"[color=ff0000]{except_obj}[/color]"
        except DuplicatedUserException as except_obj:
            self.ids.message.text = f"[color=ff0000]{except_obj}[/color]"

class MainWindow(Screen):

    def __init__(self, 
                 ui_assets_manager, 
                 message_builder_provider, 
                 profiles_manager_provider,
                 settings_provider, 
                 **kwargs):
        super(MainWindow, self).__init__(**kwargs)
        self.ui_assets_manager = ui_assets_manager
        self.message_builder_provider = message_builder_provider
        self.message_builder_provider.add_notified(self)
        self.profiles_manager_provider = profiles_manager_provider
        self.profiles_manager_provider.add_notified(self)
        self.settings_provider = settings_provider
        self.settings_provider.add_notified(self)
        
        self.settings_items_vertical_layout = None

        Window.size = (2048, 1080)

    def go_to_login_page(self):
        self.manager.current = "login_screen"

    def build(self):
        self.ui_assets_manager.set_language(self.settings_provider.get_current_language())

        # Layout principal
        self.main_layout = BoxLayout(orientation='vertical')  # Cambia la orientación a vertical
        self.add_widget(self.main_layout)
                
        # Añade un espacio en blanco 
        blank_space = BoxLayout(size_hint=(1, .03))

        # Layout de los botones
        buttons_layout = BoxLayout(orientation='horizontal', size_hint=(1, .17), spacing=10)
        self.vertical_layout = BoxLayout(orientation='vertical', size_hint=(1, 0.2))
        self.vertical_layout.add_widget(buttons_layout)
        self.main_layout.add_widget(self.vertical_layout)

        self.settings_items_vertical_layout = BoxLayout(orientation='vertical',
                                                        size_hint_x=None, 
                                                        width=200)

        # Botón para cambiar el modo de las celdas
        self.but_tile_mode = ButtonRnd(text=self.ui_assets_manager.get_string("modo_celdas"), 
                                     on_press=self.on_click_but_tile_mode, 
                                     font_name='Texto')
        self.settings_items_vertical_layout.add_widget(self.but_tile_mode)

        # Botón de idioma
        self.but_language = ButtonRnd(text=self.ui_assets_manager.get_string("idioma"), 
                            on_press=self.on_click_but_language, 
                            font_name='Texto')
        self.settings_items_vertical_layout.add_widget(self.but_language)

        current_user = self.profiles_manager_provider.get_current_user()

        self.label_username = ButtonRnd(text="", font_size=40, disabled=True)
        if (current_user!=None):
            self.label_username.text = current_user.get_username()
        self.settings_items_vertical_layout.add_widget(self.label_username)

        # Botón de cerrar sesión
        self.but_close_session = ButtonRnd(text=self.ui_assets_manager.get_string("cerrar_sesion"), 
                            on_press=self.on_click_but_close_session, 
                            font_name='Texto')
        self.settings_items_vertical_layout.add_widget(self.but_close_session)
        
        buttons_layout.add_widget(self.settings_items_vertical_layout)

        # Espacio para texto
        scroll = ScrollView(size_hint=(.4, 1), scroll_type=['bars', 'content'], bar_width=10)
        self.tf_user_aac_message = CustomTextInput(
            text="",
            # Limita el ancho del texto al ancho del widget
            #size_hint_y=None,  # Esto permitirá que el TextInput se expanda a su tamaño natural
            height=Window.height * 0.2,  # Altura inicial del TextInput
            halign='left',  
            font_name='Texto', 
            font_size=40,
            background_color=(0.7, 0.7, 0.7, 1),
            foreground_color=(0, 0, 0, 1),
            readonly=True,
            cursor_blink=False,  # Deshabilita el parpadeo del cursor
            cursor_width=0,  # Establece el ancho del cursor a 0
            focus=False
        )

        self.tf_user_aac_message.bind(on_text=self.on_tf_user_aac_message_text_changed)  # Añade un evento para cuando el texto cambie
        self.tf_user_aac_message.bind(on_touch_down = self.on_tf_user_aac_message_touch_down)
        scroll.add_widget(self.tf_user_aac_message)
        buttons_layout.add_widget(scroll)

        self.delete_buttons_vertical_layout = BoxLayout(orientation='vertical',
                                                        size_hint_x=None, 
                                                        width=200)
        # El boton para borrar una palabra
        self.but_delete_last_word = ButtonRnd(text=self.ui_assets_manager.get_string("borrar"),  
                                              on_press=self.on_click_but_delete_last_word, 
                                              font_name='Texto')
        self.delete_buttons_vertical_layout.add_widget(self.but_delete_last_word)

        # El boton para borrar todo el texto
        self.but_clear_message = ButtonRnd(text=self.ui_assets_manager.get_string("borrar_todo"), 
                                           on_press=self.on_click_but_clear_message, 
                                           font_name='Texto')
        self.delete_buttons_vertical_layout.add_widget(self.but_clear_message)

        buttons_layout.add_widget(self.delete_buttons_vertical_layout)

        self.grids_content = {}
        self.build_grid_content()

        self.grid_window = None
        self.swap_grid_window('TAB. INICIAL')

        self.vertical_layout.add_widget(blank_space)

    def build_grid_content(self):
        language = self.settings_provider.get_current_language()
        filename = self.ui_assets_manager.get_resource(f'src/grids/grids_{language}.xlsx')
        wb = load_workbook(filename)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            words_with_images = []
            for ws_row_aux in ws.iter_rows():
                current_row_content = []
                for i in range(0, len(ws_row_aux), 2):
                    tile_image = ws_row_aux[i]
                    tile_word = ws_row_aux[i + 1]
                    current_row_content.append((tile_image.value, tile_word.value))
                
                words_with_images.append(current_row_content)

            self.grids_content[sheet_name] = words_with_images

    def update_message(self):
        message = self.message_builder_provider.get_current_message()
        message_str = ""
        for current_word in message:
            message_str+=" " + current_word

        self.tf_user_aac_message.text = message_str

    def update_language(self):
        self.message_builder_provider.clear_message()
        self.build()

    def update_login(self):
        current_user = self.profiles_manager_provider.get_current_user()
        if (current_user!=None):
            self.build()
        else:
            self.go_to_login_page()

    def swap_grid_window(self, tab_id):
        if (self.grid_window is not None):
            self.main_layout.remove_widget(self.grid_window)
        
        self.grid_window = GridWindow(self, tab_id, size_hint=(1, 0.8))
        self.grid_window.set_ui_assets_manager(self.ui_assets_manager)
        self.grid_window.set_content(self.grids_content[tab_id], 
                                     self.settings_provider.get_current_tile_mode())
        self.grid_window.set_message_builder_provider(self.message_builder_provider)
        self.grid_window.build()

        self.main_layout.add_widget(self.grid_window, index=0)
        self.main_layout.do_layout()

    def on_tf_user_aac_message_text_changed(self, widget):
        pass

    def on_tf_user_aac_message_touch_down(self, widget, touch):
        pass

    def on_click_but_language(self, widget):
        available_languages = self.settings_provider.get_available_languages()
        current_language = self.settings_provider.get_current_language()
        
        for idx, language_aux in enumerate(available_languages):
            if (current_language==language_aux):
                current_language_idx = (idx+1)%(len(available_languages))

        self.settings_provider.change_current_language(available_languages[current_language_idx])

    def on_click_but_close_session(self, widget):
        self.profiles_manager_provider.logout_current_user()

    def on_click_but_tile_mode(self, widget):
        self.message_builder_provider.clear_message()
        available_tile_modes = self.settings_provider.get_available_tile_modes()
        current_tile_mode = self.settings_provider.get_current_tile_mode()

        for idx, tile_mode_aux in enumerate(available_tile_modes):
            if (current_tile_mode==tile_mode_aux):
                current_tile_mode_idx = (idx+1)%(len(available_tile_modes))
        
        self.settings_provider.change_current_tile_mode(available_tile_modes[current_tile_mode_idx])

    def on_click_but_delete_last_word(self, widget):
        self.message_builder_provider.remove_last_word_from_message()

    def on_click_but_clear_message(self, widget):
        self.message_builder_provider.clear_message()
